#!/usr/bin/env python3
"""
Build a tag map from Excel keywords to matched posts.

For each of 4952 keywords:
- Get post by slug (exact match) → 1133
- Get post by fuzzy slug match → 621
- No match → generate new slug based on keyword (will be added later when post is created)

Output: src/data/keyword-to-posts.json (map keyword.slug → [post.slug, ...])
         And adds tag field to posts.json for the matched keywords (multiple tags per post).
"""
import json
import os
import re
from openpyxl import load_workbook

EXCEL = "/Users/maabook/Desktop/beriklan.co.id/data/Master_Keyword_Plan_Beriklan.xlsx"
WEB = "/Users/maabook/Desktop/beriklan.co.id/web"
POSTS_JSON = os.path.join(WEB, "src/data/posts.json")
MAP_JSON = os.path.join(WEB, "src/data/keyword-to-posts.json")


def slugify(s):
    s = re.sub(r'[^a-z0-9]+', '-', s.lower()).strip('-')
    return s


# Load Excel
wb = load_workbook(EXCEL)
ws = wb['All Keywords (4952)']
rows = list(ws.iter_rows(min_row=2, values_only=True))

# Load posts
posts = json.load(open(POSTS_JSON))
post_by_slug = {p['slug']: p for p in posts if p.get('slug')}

# Build map
keyword_map = {}      # keyword.slug → {keyword, service, city, intent, post_slugs (list)}
unmatched_keywords = []

for row in rows:
    if not row or not row[0]: continue
    no, keyword_text, service, city, intent, slug_excel = row[:6]
    keyword_text = (keyword_text or '').strip()
    service = (service or '').strip()
    city = (city or '').strip()
    intent = (intent or '').strip()
    slug_excel = (slug_excel or '').strip()
    if not slug_excel: continue

    # Find matched posts
    matched = []
    if slug_excel in post_by_slug:
        matched = [slug_excel]
    else:
        # Fuzzy: keyword slug's last 3 tokens match a post slug
        parts = slug_excel.split('-')
        core = '-'.join(parts[-3:])
        for ps in post_by_slug:
            psp = ps.split('-')
            core_words = core.split('-')
            if len(core_words) >= 2 and all(c in psp for c in core_words):
                matched.append(ps)
                if len(matched) >= 5: break

    keyword_map[slug_excel] = {
        "keyword": keyword_text,
        "slug": slug_excel,
        "service": service,
        "city": city,
        "intent": intent,
        "post_slugs": matched[:3],  # up to 3 posts
        "has_posts": len(matched) > 0,
    }
    if not matched:
        unmatched_keywords.append(slug_excel)

# Update posts: add all unique matched tags to each post
for post in posts:
    matched_tags = set()
    for slug, info in keyword_map.items():
        if post['slug'] in info['post_slugs']:
            matched_tags.add(slug)
    if matched_tags:
        post['tags'] = list(set(post.get('tags', []) + sorted(matched_tags)))

# Save
json.dump(posts, open(POSTS_JSON, 'w'), ensure_ascii=False, indent=1)
json.dump(keyword_map, open(MAP_JSON, 'w'), ensure_ascii=False, indent=1)

print(f"total keywords: {len(keyword_map)}")
print(f"keywords with posts: {sum(1 for k in keyword_map.values() if k['has_posts'])}")
print(f"keywords without posts: {len(unmatched_keywords)}")
print(f"map saved to: {MAP_JSON}")
print(f"posts updated: {POSTS_JSON}")
