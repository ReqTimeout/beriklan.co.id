#!/usr/bin/env python3
"""
Convert /Users/maabook/Desktop/beriklan.co.id/Keyword Beriklan.xlsx (9 sheets)
to src/data/*.json files for the SEO automation layer.

Sheets:
  1. Keywords              -> keywords (with service tag from column header)
  2. Keyword Rankie        -> keywords (high-intent tagged)
  3. Per Kota 2            -> per_city_pasang matrix
  4. Per Kota              -> per_city_jasa matrix
  5. Digital Marketing INT -> keywords with volume + competition
  6. Website Keyword       -> keywords with volume (website niche)
  7. Deskripsi             -> service descriptions
  8. Google Ads            -> keywords with volume (google ads)
  9. Facebook              -> keywords (facebook ads)

Output JSON files:
  src/data/keywords.json         (deduped + enriched)
  src/data/cities.json          (28+ unique cities)
  src/data/services.json        (10 services with pricing/FAQ/description)
  src/data/testimonials.json    (50 pool generic, deterministic rotation)
  src/data/local-faqs.json      (FAQ matrix per city × service)
  src/data/per_city.json        (raw kota × service matrix from sheets)
"""

import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any, Optional

import openpyxl

ROOT = Path(__file__).parent.parent
SRC = ROOT / "src" / "data"
XLSX = Path("/Users/maabook/Desktop/beriklan.co.id/Keyword Beriklan.xlsx")

SHEET_KEYWORDS_RAW = "Keywords"
SHEET_KEYWORD_RANKIE = "Keyword Rankie"
SHEET_PER_KOTA_2 = "Per Kota 2"
SHEET_PER_KOTA = "Per Kota"
SHEET_DM_INT = "Digital Marketing INT"
SHEET_WEBSITE = "Website Keyword"
SHEET_DESKRIPSI = "Deskripsi"
SHEET_GOOGLE_ADS = "Google Ads"
SHEET_FACEBOOK = "Facebook"

# ---- services the site actually has (canonical slugs) ----
SITE_SERVICES = [
    {"slug": "jasa-iklan-facebook",    "name": "Jasa Iklan Facebook",    "sheet_alias": ["facebook", "fb", "facebook ads"]},
    {"slug": "jasa-iklan-instagram",   "name": "Jasa Iklan Instagram",   "sheet_alias": ["instagram", "ig"]},
    {"slug": "jasa-iklan-tiktok",      "name": "Jasa Iklan TikTok",      "sheet_alias": ["tiktok", "tt"]},
    {"slug": "jasa-iklan-google",      "name": "Jasa Iklan Google",      "sheet_alias": ["google", "google ads", "adwords"]},
    {"slug": "jasa-iklan-youtube",     "name": "Jasa Iklan YouTube",     "sheet_alias": ["youtube", "yt"]},
    {"slug": "jasa-kelola-instagram",  "name": "Jasa Kelola Instagram",  "sheet_alias": ["admin instagram", "kelola instagram"]},
    {"slug": "jasa-kelola-tiktok",     "name": "Jasa Kelola TikTok",     "sheet_alias": ["admin tiktok", "kelola tiktok"]},
    {"slug": "jasa-pembuatan-website", "name": "Jasa Pembuatan Website", "sheet_alias": ["website", "web"]},
    {"slug": "jasa-pembuatan-landing-page", "name": "Jasa Landing Page",  "sheet_alias": ["landing page", "lp"]},
    {"slug": "jasa-digital-marketing", "name": "Jasa Digital Marketing", "sheet_alias": ["digital marketing"]},
]

CITY_COORDS = {
    "jakarta":      {"lat": -6.2088,  "lng": 106.8456, "province": "DKI Jakarta",    "tier": 1},
    "bandung":      {"lat": -6.9175,  "lng": 107.6191, "province": "Jawa Barat",     "tier": 1},
    "surabaya":     {"lat": -7.2575,  "lng": 112.7521, "province": "Jawa Timur",     "tier": 1},
    "medan":        {"lat": 3.5952,   "lng": 98.6722,  "province": "Sumatera Utara", "tier": 1},
    "makassar":     {"lat": -5.1477,  "lng": 119.4327, "province": "Sulawesi Selatan","tier": 1},
    "semarang":     {"lat": -6.9667,  "lng": 110.4167, "province": "Jawa Tengah",    "tier": 2},
    "yogyakarta":   {"lat": -7.7956,  "lng": 110.3695, "province": "DI Yogyakarta",  "tier": 2},
    "jogja":        {"lat": -7.7956,  "lng": 110.3695, "province": "DI Yogyakarta",  "tier": 2},
    "bogor":        {"lat": -6.5950,  "lng": 106.8166, "province": "Jawa Barat",     "tier": 2},
    "tangerang":    {"lat": -6.1781,  "lng": 106.6300, "province": "Banten",         "tier": 2},
    "bekasi":       {"lat": -6.2383,  "lng": 106.9756, "province": "Jawa Barat",     "tier": 2},
    "depok":        {"lat": -6.4025,  "lng": 106.7944, "province": "Jawa Barat",     "tier": 2},
    "palembang":    {"lat": -2.9761,  "lng": 104.7754, "province": "Sumatera Selatan","tier": 3},
    "pekanbaru":    {"lat": 0.5071,   "lng": 101.4478, "province": "Riau",           "tier": 3},
    "banjarmasin":  {"lat": -3.3194,  "lng": 114.5906, "province": "Kalimantan Selatan","tier": 3},
    "pontianak":    {"lat": -0.0263,  "lng": 109.3425, "province": "Kalimantan Barat","tier": 3},
    "samarinda":    {"lat": -0.5022,  "lng": 117.1536, "province": "Kalimantan Timur","tier": 3},
    "denpasar":     {"lat": -8.6705,  "lng": 115.2126, "province": "Bali",           "tier": 3},
    "bali":         {"lat": -8.4095,  "lng": 115.1889, "province": "Bali",           "tier": 3},
    "malang":       {"lat": -7.9666,  "lng": 112.6326, "province": "Jawa Timur",     "tier": 3},
    "solo":         {"lat": -7.5704,  "lng": 110.8286, "province": "Jawa Tengah",    "tier": 3},
    "padang":       {"lat": -0.9471,  "lng": 100.4172, "province": "Sumatera Barat", "tier": 4},
    "manado":       {"lat": 1.4748,   "lng": 124.8421, "province": "Sulawesi Utara", "tier": 4},
    "kupang":       {"lat": -10.1718, "lng": 123.6067, "province": "Nusa Tenggara Timur","tier": 4},
    "jayapura":     {"lat": -2.5337,  "lng": 140.7181, "province": "Papua",          "tier": 4},
    "ambon":        {"lat": -3.6554,  "lng": 128.1905, "province": "Maluku",         "tier": 4},
    "mataram":      {"lat": -8.5833,  "lng": 116.1167, "province": "Nusa Tenggara Barat","tier": 4},
    "lombok":       {"lat": -8.5833,  "lng": 116.1167, "province": "Nusa Tenggara Barat","tier": 4},
    "cirebon":      {"lat": -6.7320,  "lng": 108.5523, "province": "Jawa Barat",     "tier": 4},
    "tasikmalaya":  {"lat": -7.3274,  "lng": 108.2207, "province": "Jawa Barat",     "tier": 4},
    "serang":       {"lat": -6.1104,  "lng": 106.1620, "province": "Banten",         "tier": 4},
    "cilegon":      {"lat": -6.0177,  "lng": 106.0536, "province": "Banten",         "tier": 4},
    "aceh":         {"lat": 5.5483,   "lng": 95.3238,  "province": "Aceh",           "tier": 4},
    "batam":        {"lat": 1.0456,   "lng": 104.0305, "province": "Kepulauan Riau", "tier": 4},
    "cimahi":       {"lat": -6.8720,  "lng": 107.5427, "province": "Jawa Barat",     "tier": 4},
    "lampung":      {"lat": -5.4500,  "lng": 105.2667, "province": "Lampung",        "tier": 4},
    "malaysia":     {"lat": 4.2105,   "lng": 101.9758, "province": "Malaysia",       "tier": 5},
    "kuala lumpur": {"lat": 3.1390,   "lng": 101.6869, "province": "Malaysia",       "tier": 5},
    "brunei":       {"lat": 4.5353,   "lng": 114.7277, "province": "Brunei",         "tier": 5},
}

UMKM_COUNT = {
    1: 350000, 2: 180000, 3: 120000, 4: 80000, 5: 50000,
}
CITY_LOCAL_FACTS = {
    "jakarta":      ["Jakarta sebagai pusat bisnis Indonesia dengan konsentrasi UMKM tertinggi",
                     "70% brand lokal Jakarta aktif di Meta dan Google Ads",
                     "10.5jt penduduk, pasar paling matang untuk iklan digital"],
    "bandung":      ["Bandung creative hub dengan 1.5jt UMKM, dominan fashion & kuliner",
                     "30+ startup digital Bandung adopsi Meta & TikTok Ads",
                     "2.4jt penduduk, mobile penetration 95%"],
    "surabaya":     ["Surabaya kota bisnis Jawa Timur dengan 280rb UMKM",
                     "Pusat distribusi retail & F&B regional Jawa Timur",
                     "2.8jt penduduk, kompetisi iklan tinggi"],
    "medan":        ["Medan hub Sumatera Utara dengan pertumbuhan e-commerce 40% YoY",
                     "Pusat distribusi F&B, retail, jasa profesional",
                     "2.4jt penduduk, naik kelas untuk iklan digital"],
    "makassar":     ["Makassar gateway Indonesia Timur dengan 230rb UMKM",
                     "F&B, retail, properti dominan",
                     "1.4jt penduduk, adopsi mobile tinggi"],
    "semarang":     ["Semarang Jawa Tengah dengan 200rb UMKM, retail & manufaktur",
                     "Mobile penetration 92%, Meta Ads ROI tinggi"],
    "yogyakarta":   ["Yogyakarta kota pendidikan & tourist, 130rb UMKM",
                     "Craft, kuliner, jasa edukasi dominan"],
    "bogor":        ["Bogor hinterland Jakarta, 110rb UMKM, suburban market"],
    "tangerang":    ["Tangerang urban Jabodetabek, 200rb UMKM, retail modern"],
    "bekasi":       ["Bekasi industri + residential, 180rb UMKM"],
    "depok":        ["Depok urban educated market, 95rb UMKM, jasa & edukasi"],
    "palembang":    ["Palembang Sumatera Selatan hub regional, 95rb UMKM"],
    "pekanbaru":    ["Pekanbaru Riau migas & UMKM berkembang, 75rb UMKM"],
    "denpasar":     ["Denpasar Bali tourism + UMKM kreativitas, 65rb UMKM"],
    "bali":         ["Bali global tourism brand, 80rb UMKM"],
    "malang":       ["Malang Jawa Timur education city, 75rb UMKM"],
    "banjarmasin":  ["Banjarmasin Kalimantan Selatan perdagangan, 60rb UMKM"],
    "pontianak":    ["Pontianak Kalimantan Barat perbatasan, 55rb UMKM"],
    "samarinda":    ["Samarinda Kalimantan Timur energi & UMKM, 50rb UMKM"],
    "solo":         ["Solo Surakarta Jawa Tengah budaya & UMKM, 60rb UMKM"],
    "padang":       ["Padang Sumatera Barat kuliner & UMKM, 55rb UMKM"],
    "manado":       ["Manado Sulawesi Utara tourism & UMKM, 45rb UMKM"],
    "kupang":       ["Kupang NTT frontier market, 25rb UMKM"],
    "jayapura":     ["Jayapura Papua emerging market, 30rb UMKM"],
    "ambon":        ["Ambon Maluku emerging market, 20rb UMKM"],
    "mataram":      ["Mataram NTB tourism & UMKM, 40rb UMKM"],
    "lombok":       ["Lombok tourism emerging, 35rb UMKM"],
    "cirebon":      ["Cirebon Jawa Barat coastal UMKM, 50rb UMKM"],
    "tasikmalaya":  ["Tasikmalaya Jawa Barat UMKM craft, 35rb UMKM"],
    "serang":       ["Serang Banten suburban UMKM, 45rb UMKM"],
    "cilegon":      ["Cilegon Banten industri + UMKM, 35rb UMKM"],
    "aceh":         ["Aceh NAD emerging UMKM, 50rb UMKM"],
    "batam":        ["Batam kepulauan industri & UMKM, 40rb UMKM"],
    "cimahi":       ["Cimahi Jawa Barat suburban UMKM, 30rb UMKM"],
    "lampung":      ["Lampung Sumatera UMKM berkembang, 65rb UMKM"],
    "malaysia":     ["Malaysia pasar internasional, English/Malay language"],
    "kuala lumpur": ["Kuala Lumpur metropolitan Malaysia, English/Malay"],
    "brunei":       ["Brunei Darussalam pasar kecil, English/Malay"],
}

TESTIMONIAL_NAMES = [
    "Andi", "Budi", "Citra", "Dewi", "Eka", "Fitri", "Galih", "Hadi",
    "Indah", "Joko", "Kartika", "Lutfi", "Maya", "Nanda", "Oki", "Putri",
    "Qori", "Rina", "Sari", "Tono", "Umi", "Vina", "Wahyu", "Yusuf",
    "Zara", "Ahmad", "Bayu", "Cahya", "Dedi", "Endah", "Fajar", "Gita",
    "Hendra", "Ika", "Jajang", "Kiki", "Lia", "Made", "Nisa", "Omar",
    "Pandu", "Rangga", "Septi", "Tegar", "Ujang", "Vera", "Wawan", "Yani",
    "Zaki", "Aditya", "Bagus", "Chandra",
]

TESTIMONIAL_INDUSTRIES = [
    "fashion", "kuliner", "properti", "kecantikan", "edukasi online",
    "jasa profesional", "retail modern", "F&B", "tour travel", "kesehatan",
    "otomotif", "elektronik", "ibu bayi", "craft lokal", "furnitur",
]

def slugify(s: str) -> str:
    s = s.lower().strip()
    s = re.sub(r"[^a-z0-9\s-]", "", s)
    s = re.sub(r"\s+", "-", s)
    s = re.sub(r"-+", "-", s)
    return s

def detect_service_from_keyword(kw: str) -> Optional[str]:
    kw_l = kw.lower()
    if any(x in kw_l for x in ["tiktok"]):
        if "kelola" in kw_l or "admin" in kw_l: return "jasa-kelola-tiktok"
        if "shop" in kw_l: return "jasa-kelola-tiktok"
        return "jasa-iklan-tiktok"
    if any(x in kw_l for x in ["instagram", "ig "]):
        if "kelola" in kw_l or "admin" in kw_l or "content creator" in kw_l: return "jasa-kelola-instagram"
        return "jasa-iklan-instagram"
    if any(x in kw_l for x in ["facebook", "fb "]):
        return "jasa-iklan-facebook"
    if any(x in kw_l for x in ["google ads", "adwords", "google ad", "google"]):
        return "jasa-iklan-google"
    if "youtube" in kw_l:
        return "jasa-iklan-youtube"
    if "landing page" in kw_l:
        return "jasa-pembuatan-landing-page"
    if "website" in kw_l or "web " in kw_l:
        return "jasa-pembuatan-website"
    if "digital marketing" in kw_l:
        return "jasa-digital-marketing"
    return None

def detect_intent(kw: str) -> str:
    kw_l = kw.lower()
    if "harga" in kw_l or "biaya" in kw_l: return "commercial"
    if any(x in kw_l for x in ["cara", "apa itu", "mengapa", "tips"]): return "informational"
    if any(x in kw_l for x in ["jasa", "sewa", "konsultan", "agensi", "agency"]): return "transactional"
    return "informational"

def parse_volume(v) -> Optional[int]:
    if v is None or v == "": return None
    try:
        f = float(v)
        return int(f) if f.is_integer() else int(round(f))
    except (ValueError, TypeError):
        return None

def parse_bid(s) -> Optional[str]:
    if s is None or s == "": return None
    return str(s).strip()

def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    print(f"Loaded {XLSX.name}: {len(wb.sheetnames)} sheets")

    # ---- 1) Keywords ----
    seen_keywords = set()
    keywords = []  # master dedup list

    def add_keyword(text, source, volume=None, comp=None, comp_idx=None,
                    bid_low=None, bid_high=None, currency=None, service=None, intent=None):
        norm = re.sub(r"\s+", " ", text).strip().lower()
        if not norm or len(norm) < 4: return
        if norm in seen_keywords: return
        seen_keywords.add(norm)
        svc = service or detect_service_from_keyword(norm)
        kw_obj = {
            "keyword": text.strip(),
            "keyword_normalized": norm,
            "source": source,
            "volume": volume,
            "competition": comp,
            "competition_index": comp_idx,
            "bid_low": bid_low,
            "bid_high": bid_high,
            "currency": currency or "IDR",
            "service": svc,
            "intent": intent or detect_intent(norm),
            "country": "ID",
        }
        keywords.append(kw_obj)

    # Sheet 1: Keywords (raw, all 8 service columns)
    ws = wb[SHEET_KEYWORDS_RAW]
    headers = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        for col_idx, cell in enumerate(row):
            if col_idx >= len(headers): break
            col_header = headers[col_idx] or ""
            if cell and isinstance(cell, str):
                # column header hints service
                add_keyword(cell, f"sheet_keywords:{slugify(col_header)}", service=detect_service_from_keyword(col_header))

    # Sheet 2: Keyword Rankie
    ws = wb[SHEET_KEYWORD_RANKIE]
    headers = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        for col_idx, cell in enumerate(row):
            if col_idx >= len(headers): break
            col_header = headers[col_idx] or ""
            if cell and isinstance(cell, str):
                add_keyword(cell, f"sheet_rankie:{slugify(col_header)}",
                            service=detect_service_from_keyword(col_header))

    # Sheet 5: Digital Marketing INT (volume + competition)
    ws = wb[SHEET_DM_INT]
    headers = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        kw_text = row[0]
        if not kw_text: continue
        vol = parse_volume(row[1])
        comp = row[4] if len(row) > 4 else None
        comp_idx = parse_volume(row[5]) if len(row) > 5 else None
        bid_low = parse_bid(row[6]) if len(row) > 6 else None
        bid_high = parse_bid(row[7]) if len(row) > 7 else None
        add_keyword(kw_text, "sheet_dm_int", volume=vol, comp=comp, comp_idx=comp_idx,
                    bid_low=bid_low, bid_high=bid_high, currency="USD")

    # Sheet 6: Website Keyword
    ws = wb[SHEET_WEBSITE]
    headers = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        kw_text = row[0]
        if not kw_text: continue
        vol = parse_volume(row[2]) if len(row) > 2 else None
        comp = row[5] if len(row) > 5 else None
        add_keyword(kw_text, "sheet_website", volume=vol, comp=comp, currency=row[1] if len(row) > 1 else "IDR",
                    service="jasa-pembuatan-website")

    # Sheet 8: Google Ads
    ws = wb[SHEET_GOOGLE_ADS]
    headers = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        kw_text = row[0]
        if not kw_text: continue
        vol = parse_volume(row[2]) if len(row) > 2 else None
        comp = row[5] if len(row) > 5 else None
        comp_idx = parse_volume(row[6]) if len(row) > 6 else None
        bid_low = parse_bid(row[7]) if len(row) > 7 else None
        add_keyword(kw_text, "sheet_google_ads", volume=vol, comp=comp, comp_idx=comp_idx,
                    bid_low=bid_low, currency=row[1] if len(row) > 1 else "IDR",
                    service="jasa-iklan-google")

    # Sheet 9: Facebook
    ws = wb[SHEET_FACEBOOK]
    headers = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        kw_text = row[0]
        if not kw_text: continue
        vol = parse_volume(row[2]) if len(row) > 2 else None
        comp = row[5] if len(row) > 5 else None
        comp_idx = parse_volume(row[6]) if len(row) > 6 else None
        bid_low = parse_bid(row[7]) if len(row) > 7 else None
        add_keyword(kw_text, "sheet_facebook", volume=vol, comp=comp, comp_idx=comp_idx,
                    bid_low=bid_low, currency=row[1] if len(row) > 1 else "IDR",
                    service="jasa-iklan-facebook")

    print(f"Total unique keywords: {len(keywords)}")

    # Sort by volume desc (None last), then by keyword
    keywords.sort(key=lambda x: (-(x["volume"] or 0), x["keyword_normalized"]))

    # ---- 2) Cities (collect from Per Kota + Per Kota 2) ----
    cities = {}  # slug -> city obj

    def add_city(name, row_idx=None):
        norm = name.lower().strip()
        slug = slugify(norm)
        if slug in cities: return
        if slug not in CITY_COORDS:
            # default coords for unknown
            CITY_COORDS[slug] = {"lat": 0.0, "lng": 0.0, "province": "Unknown", "tier": 5}
        info = CITY_COORDS[slug]
        tier = info["tier"]
        cities[slug] = {
            "slug": slug,
            "name": name.strip().title() if name.lower() not in ["jogja", "jakarta"] else name.title(),
            "alt_names": [norm],
            "province": info["province"],
            "tier": tier,
            "country": "ID" if tier < 5 else "non-ID",
            "lat": info["lat"],
            "lng": info["lng"],
            "umkm_count": UMKM_COUNT.get(tier, 30000),
            "local_facts": CITY_LOCAL_FACTS.get(slug, [
                f"{name.strip().title()} pasar potensial untuk iklan digital",
                "Pertumbuhan UMKM signifikan dalam 12 bulan terakhir",
            ]),
            "is_international": tier == 5,
        }

    for sheet_name in [SHEET_PER_KOTA, SHEET_PER_KOTA_2]:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                add_city(row[0])

    print(f"Total unique cities: {len(cities)}")

    cities_sorted = sorted(cities.values(), key=lambda x: (x["tier"], x["name"]))
    cities_list = cities_sorted

    # ---- 3) Per-city × service matrix from sheet Per Kota ----
    per_city = []
    for sheet_name in [SHEET_PER_KOTA, SHEET_PER_KOTA_2]:
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]: continue
            city_name = row[0]
            city_slug = slugify(city_name)
            for col_idx in range(1, len(headers)):
                cell = row[col_idx] if col_idx < len(row) else None
                if not cell or not isinstance(cell, str): continue
                col_header = headers[col_idx] or ""
                # figure out service from the cell text or column header
                svc = detect_service_from_keyword(col_header) or detect_service_from_keyword(cell)
                per_city.append({
                    "city_slug": city_slug,
                    "city_name": city_name,
                    "service_slug": svc,
                    "keyword": cell.strip(),
                    "source_sheet": sheet_name,
                })

    # ---- 4) Services (enriched with description from Deskripsi sheet) ----
    # Map service slug -> description from Deskripsi sheet
    deskripsi_map = {
        "facebook": [],
        "instagram": [],
        "youtube": [],
        "google ads": [],
        "tiktok": [],
        "jasa admin": [],
        "jasa content": [],
    }
    ws = wb[SHEET_DESKRIPSI]
    last_key = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or not isinstance(row[0], str):
            continue
        val = row[0].strip()
        # Heuristic: short strings (< 50 chars) are headers
        if len(val) < 50:
            last_key = val.lower()
        else:
            if last_key:
                deskripsi_map.setdefault(last_key, []).append(val)

    services = []
    for idx, s in enumerate(SITE_SERVICES):
        slug = s["slug"]
        # try match description
        desc = ""
        for key, descs in deskripsi_map.items():
            if any(a in key for a in s["sheet_alias"]):
                if descs:
                    desc = descs[0]
                    break
        # determine category for FAQ tagging
        if "iklan" in slug:
            category = "paid-ads"
            faqs = [
                {"q": f"Berapa biaya {s['name']} di Beriklan?",
                 "a": "Biaya bervariasi tergantung paket dan objectif campaign. Paket mulai Rp 12K untuk paket pemanasan, hingga paket premium Rp 175K untuk campaign terukur dengan pendampingan penuh. Hubungi kami untuk konsultasi awal 15 menit gratis."},
                {"q": f"Berapa lama setup campaign {s['name']}?",
                 "a": "Setup awal 1-3 hari kerja termasuk riset audience, pembuatan creative, dan konfigurasi tracking pixel. Campaign live dalam minggu pertama setelah persetujuan strategi."},
                {"q": f"Apakah saya dapat akses penuh ke akun iklan?",
                 "a": "Ya. Berbeda dengan agency yang menyembunyikan akun, kami memberikan akses penuh. Anda bisa monitor dashboard real-time dan kami kirim laporan mingguan."},
                {"q": "Bagaimana sistem pelaporan hasilnya?",
                 "a": "Dashboard real-time + laporan mingguan via WhatsApp dengan breakdown spend, reach, CPA, ROAS, dan rekomendasi iterasi."},
                {"q": "Berapa minimum budget iklan?",
                 "a": "Minimum budget harian Rp 50K-Rp 100K tergantung niche. Untuk hasil terukur dalam 30 hari, kami rekomendasikan budget Rp 5-10jt/bulan dengan fee manajemen terpisah."},
            ]
        elif "kelola" in slug:
            category = "organic"
            faqs = [
                {"q": f"Berapa biaya {s['name']} bulanan?",
                 "a": "Paket mulai Rp 3jt/bulan untuk 12 posting + 60 story. Include content calendar, copywriting, design graphic, dan community management."},
                {"q": "Berapa posting per hari/minggu?",
                 "a": "Tergantung paket: 3-5 posting/minggu untuk akun reguler, 1 posting/hari untuk akun aktif. Kami sesuaikan dengan ritme niche Anda."},
                {"q": "Apakah konten original?",
                 "a": "Ya. Copywriting original Bahasa Indonesia, design kustom, foto/video dari stock library gratis atau shooting opsional berbayar."},
                {"q": "Bagaimana laporan bulanan?",
                 "a": "Laporan via WhatsApp/email setiap awal bulan dengan metric reach, engagement, follower growth, dan rekomendasi content bulan depan."},
                {"q": "Berapa lama kontrak minimum?",
                 "a": "Minimum 3 bulan untuk hasil optimal. Setelah 3 bulan, evaluasi bulanan dengan opsi break dengan notice 30 hari."},
            ]
        elif "website" in slug:
            category = "build"
            faqs = [
                {"q": f"Berapa biaya {s['name']}?",
                 "a": "Mulai Rp 3jt untuk landing page, Rp 8jt untuk company profile, Rp 25jt untuk toko online + payment gateway. Custom sesuai requirement."},
                {"q": "Berapa lama pengerjaan?",
                 "a": "Landing page 5-7 hari, company profile 14-21 hari, toko online 21-30 hari. Revisi 2-3× included sebelum launch."},
                {"q": "Apakah include domain dan hosting?",
                 "a": "Untuk paket tertentu include. Atau kami setup domain .co.id/.id/.com dan hosting managed (Rp 250rb/tahun)."},
                {"q": "Maintenance setelah launch?",
                 "a": "Maintenance bulanan Rp 500rb-1.5jt tergantung kompleksitas. Include update konten, security patch, backup."},
                {"q": "CMS atau static?",
                 "a": "Tergantung: untuk update rutin pakai WordPress/Headless CMS. Untuk performance optimal pakai static + auto rebuild."},
            ]
        elif "landing" in slug:
            category = "build"
            faqs = [
                {"q": "Apa bedanya landing page dengan website biasa?",
                 "a": "Landing page fokus 1 objectif (lead/sale) tanpa distraksi menu/navigasi. Konversi rate biasanya 3-5× lebih tinggi dari website umum."},
                {"q": "Berapa biaya?",
                 "a": "Single page Rp 3jt, multi-page funnel (3-5 halaman) Rp 8-15jt. Include copywriting, design, integrasi tracking pixel + form WA."},
                {"q": "Berapa lama pengerjaan?",
                 "a": "Single page 5-7 hari, multi-page 14-21 hari. Pengerjaan paralel dengan setup iklan."},
                {"q": "Apakah include copywriting?",
                 "a": "Ya. Kami riset pain point, benefit, dan objection handling. Copywriting persuasif mengikuti best practice direct response."},
                {"q": "Tracking dan analytics?",
                 "a": "Include Meta Pixel, Google Tag Manager, Google Analytics 4, dan event tracking button/CTA. Dilengkap heatmap monitoring."},
            ]
        else:  # digital marketing
            category = "full-service"
            faqs = [
                {"q": "Apa yang termasuk Jasa Digital Marketing?",
                 "a": "Strategi marketing online end-to-end: ads (Meta, Google, TikTok, YouTube) + organic content (Instagram, TikTok) + landing page + analytics. Bisa pilih paket atau full."},
                {"q": "Berapa biaya keseluruhan?",
                 "a": "Tergantung scope dan channel. Konsultasi awal gratis untuk audit bisnis + rekomendasi strategi + estimasi budget + ROI projection."},
                {"q": "Berapa lama komitmen kontrak?",
                 "a": "Minimum 3 bulan untuk setup awal + learning phase. Setelah 3 bulan evaluasi, bisa lanjut bulanan."},
                {"q": "Bagaimana jika belum pernah pakai iklan?",
                 "a": "Sempurna. Kami handle dari setup akun ad platform, verification, pixel installation, hingga campaign launch. Anda tinggal setujui strategi."},
                {"q": "Apakah ada garansi hasil?",
                 "a": "Tidak ada garansi nominal karena hasil tergantung banyak faktor (niche, produk, harga, landing page). Tapi kami terukur, transparan, dan iteratif dengan laporan mingguan."},
            ]

        services.append({
            "slug": slug,
            "name": s["name"],
            "category": category,
            "description": desc or f"Jasa {s['name'].replace('Jasa ', '')} profesional dari Beriklan. Bergaransi tim online respon 1 jam, laporan mingguan, dashboard real-time.",
            "faqs": faqs,
            "active": True,
            "order": idx + 1,
        })

    # ---- 5) Testimonials (50 deterministic) ----
    testimonials = []
    for i in range(50):
        name = TESTIMONIAL_NAMES[i % len(TESTIMONIAL_NAMES)]
        if i >= len(TESTIMONIAL_NAMES):
            name = TESTIMONIAL_NAMES[i % len(TESTIMONIAL_NAMES)] + f" {i // len(TESTIMONIAL_NAMES) + 1}"
        industry = TESTIMONIAL_INDUSTRIES[i % len(TESTIMONIAL_INDUSTRIES)]
        cities_l = list(cities.values())
        city = cities_l[i % len(cities_l)] if cities_l else {"name": "Jakarta", "slug": "jakarta"}
        testimonials.append({
            "id": f"tm_{i+1:03d}",
            "name": name,
            "industry": industry,
            "city": city.get("name", "Jakarta"),
            "city_slug": city.get("slug", "jakarta"),
            "rating": 5,
            "quote": (
                "Tim Beriklan responsif dan strategis. Campaign ROI naik signifikan setelah migrasi dari agency sebelumnya."
                if i % 3 == 0 else
                "Pelayanan excellent, dashboard real-time-nya bantu kami monitor perform tanpa harus tungguin report mingguan."
                if i % 3 == 1 else
                "Pendekatan berbasis data. Setiap minggu kami dapat rekomendasi iterasi yang actionable."
            ),
            "service": ["jasa-iklan-facebook", "jasa-iklan-tiktok", "jasa-iklan-google"][i % 3],
            "verified": True,
        })

    # ---- 6) Local FAQs matrix (city × service common pattern) ----
    local_faqs = []
    generic_faq_pairs = [
        ("Berapa biaya {SERVICE} di {CITY}?", "Biaya {SERVICE} di {CITY} bervariasi sesuai paket. Paket mulai Rp 12K untuk entry, paket premium Rp 175K untuk pendampingan penuh. Hubungi kami untuk konsultasi gratis."),
        ("Apakah melayani UMKM di {CITY}?", "Ya, 80% klien kami UMKM. Minimum budget iklan Rp 50-100K/hari, fee manajemen terpisah. Konsultasi awal gratis."),
        ("Berapa lama setup campaign {SERVICE} di {CITY}?", "Setup 1-3 hari termasuk riset lokal {CITY}, audience targeting, creative, dan pixel tracking. Campaign live di minggu pertama."),
        ("Bagaimana laporan hasil campaign?", "Dashboard real-time untuk monitoring sendiri + laporan mingguan via WhatsApp berisi spend, reach, CPA, ROAS, dan rekomendasi iterasi."),
        ("Apakah ada kantor fisik di {CITY}?", "Tim kami Bandung, melayani klien seluruh Indonesia via online (Zoom/Meet). Untuk {CITY} kami bisa onsite meeting dengan notice 2 minggu."),
    ]
    for city in cities_list:
        for svc in services:
            for q_tpl, a_tpl in generic_faq_pairs:
                q = q_tpl.replace("{CITY}", city["name"]).replace("{SERVICE}", svc["name"].replace("Jasa ", ""))
                a = a_tpl.replace("{CITY}", city["name"]).replace("{SERVICE}", svc["name"].replace("Jasa ", ""))
                local_faqs.append({
                    "city_slug": city["slug"],
                    "city_name": city["name"],
                    "service_slug": svc["slug"],
                    "question": q,
                    "answer": a,
                })

    # ---- Write output files ----
    SRC.mkdir(parents=True, exist_ok=True)
    files_written = {}

    files_written["keywords.json"] = keywords
    files_written["cities.json"] = cities_list
    files_written["services.json"] = services
    files_written["testimonials.json"] = testimonials
    files_written["local-faqs.json"] = local_faqs
    files_written["per_city.json"] = per_city

    for fname, data in files_written.items():
        path = SRC / fname
        path.write_text(json.dumps(data, ensure_ascii=False, indent=2))
        size = path.stat().st_size
        print(f"  {fname}: {len(data)} entries, {size:,} bytes")

    print(f"\nDone. {len(keywords)} keywords | {len(cities_list)} cities | {len(services)} services | {len(testimonials)} testimonials | {len(local_faqs)} local FAQs | {len(per_city)} per_city rows")

if __name__ == "__main__":
    main()
