#!/usr/bin/env python3
"""
Keyword Miner — Beriklan.co.id
================================
Discovers new keyword opportunities by combining:
  1. Existing keywords.json (1478 base keywords)
  2. Master_Keyword_Plan (Excel) — full keyword plan per service x city
  3. Google Suggest API (public autocomplete, no key required) — expansion per city
  4. Heuristic variants per (service, city) pair

Outputs:
  - keyword-queue.json: ranked keywords that DO NOT YET have a post published
    (target: 1000+ untapped keywords from 25 cities x 10 services)
  - keyword-queue-stats.json: mining summary

Usage:
  python3 scripts/keyword_miner.py --limit 200       # mine 200 new keywords
  python3 scripts/keyword_miner.py --all              # mine all (slow)
  python3 scripts/keyword_miner.py --dry-run          # preview only
"""
import os, sys, json, re, time, argparse, urllib.request, urllib.parse, urllib.error
from datetime import datetime, timezone

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WEB = os.path.join(ROOT, 'web')
KEYWORDS_JSON = os.path.join(WEB, 'src/data/keywords.json')
POSTS_JSON = os.path.join(WEB, 'src/data/posts.json')
KEYWORD_TO_POSTS = os.path.join(WEB, 'src/data/keyword-to-posts.json')
EXCEL_PATH = os.path.join(ROOT, 'Keyword Beriklan.xlsx')
QUEUE_OUT = os.path.join(WEB, 'src/data/keyword-queue.json')
STATS_OUT = os.path.join(WEB, 'src/data/keyword-queue-stats.json')

# Service slugs (must match /jasa-*/ page names)
SERVICES = [
    ('jasa-digital-marketing', 'digital marketing'),
    ('jasa-iklan-facebook', 'facebook ads'),
    ('jasa-iklan-instagram', 'instagram ads'),
    ('jasa-iklan-tiktok', 'tiktok ads'),
    ('jasa-iklan-google', 'google ads'),
    ('jasa-iklan-youtube', 'youtube ads'),
    ('jasa-kelola-instagram', 'kelola instagram'),
    ('jasa-kelola-tiktok', 'kelola tiktok'),
    ('jasa-pembuatan-website', 'bikin website'),
    ('jasa-pembuatan-landing-page', 'landing page'),
    ('jasa-view-live-tiktok', 'view live tiktok'),
    ('jasa-view-live-tiktok', 'live streaming tiktok'),
    ('jasa-view-live-tiktok', 'viewers tiktok'),
    ('jasa-shopee-affiliate', 'shopee affiliate'),
    ('jasa-shopee-affiliate', 'jualan shopee'),
    ('jasa-shopee-affiliate', 'shopee store'),
    ('jasa-tokopedia-affiliate', 'tokopedia affiliate'),
    ('jasa-tokopedia-affiliate', 'jualan tokopedia'),
]

# Top 25 Indonesian cities (matches the existing /jasa-*/[city]/ pages)
CITIES = [
    'jakarta', 'bandung', 'surabaya', 'medan', 'semarang',
    'makassar', 'palembang', 'tangerang', 'depok', 'bekasi',
    'bogor', 'batam', 'pekanbaru', 'denpasar', 'yogyakarta',
    'malang', 'solo', 'balikpapan', 'samarinda', 'pontianak',
    'banjarmasin', 'manado', 'padang', 'palembang', 'sidoarjo',
]

# Generic modifier templates (long-tail expansion)
LONG_TAIL_TEMPLATES = [
    '{kw} untuk pemula',
    '{kw} untuk umkm',
    '{kw} yang bagus',
    '{kw} terbaik',
    '{kw} profesional',
    '{kw} murah',
    '{kw} berkualitas',
    '{kw} berpengalaman',
    '{kw} terdekat',
    '{kw} halal',
    '{kw} gratis konsultasi',
    '{kw} harga',
    '{kw} biaya',
    '{kw} paket',
    '{kw} 2026',
    'cara {kw}',
    'tips {kw}',
    'belajar {kw}',
    'manfaat {kw}',
    'konsultan {kw}',
]

EXCLUDE_NICHE = re.compile(
    r'\b(mesum|seks|porn|casino|judi|togel|slot|hack|crack|free download|'
    r'cheat|bokep|ml|only fans|bitcoin|crypto trading)\b',
    re.IGNORECASE,
)


def load_existing_keywords():
    """Load all existing keywords + slugs that already have posts."""
    keywords = json.load(open(KEYWORDS_JSON))
    posts = json.load(open(POSTS_JSON))
    slug_set = {p['slug'] for p in posts}
    print(f'  Loaded {len(keywords)} keywords, {len(posts)} posts', file=sys.stderr)
    return keywords, slug_set


def keyword_to_slug(kw):
    """Convert keyword text to slug (matches Astro route format)."""
    s = kw.lower().strip()
    s = re.sub(r'[^a-z0-9\s-]', '', s)
    s = re.sub(r'\s+', '-', s)
    s = re.sub(r'-+', '-', s).strip('-')
    return s


def google_suggest(query, max_results=8):
    """Fetch Google Suggest autocomplete for a query. Returns list of strings."""
    try:
        url = 'http://suggestqueries.google.com/complete/search?client=firefox&q=' + urllib.parse.quote(query)
        req = urllib.request.Request(url, headers={
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
        })
        with urllib.request.urlopen(req, timeout=10) as r:
            data = json.loads(r.read().decode())
            return data[1][:max_results] if len(data) > 1 else []
    except Exception as e:
        return []


def mine_from_excel():
    """Read keywords from Master_Keyword_Plan Excel."""
    if not os.path.exists(EXCEL_PATH):
        return []
    try:
        import openpyxl
    except ImportError:
        return []
    keywords = []
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str) and len(cell) > 5 and len(cell) < 80:
                        if not EXCLUDE_NICHE.search(cell):
                            keywords.append(cell.strip())
    except Exception as e:
        print(f'  WARN: Excel read failed: {e}', file=sys.stderr)
    return keywords


def mine_from_suggestions():
    """For each (service, city) pair, query Google Suggest with multiple variants.
    Parallelized with ThreadPoolExecutor (10 workers) for speed.
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed

    # Build all query tasks: (service_idx, city, query_text)
    tasks = []
    for service_slug, service_name in SERVICES:
        for city in CITIES:
            base_queries = [
                f'{service_name} {city}',
                f'jasa {service_name} {city}',
            ]
            for q in base_queries:
                tasks.append(q)

    print(f'   launching {len(tasks)} Suggest queries in parallel (10 workers)...', file=sys.stderr)
    results = []
    with ThreadPoolExecutor(max_workers=10) as ex:
        futures = {ex.submit(google_suggest, q, 4): q for q in tasks}
        done = 0
        for f in as_completed(futures):
            done += 1
            sugg = f.result()
            for s in sugg:
                s = s.strip()
                if 5 < len(s) < 80 and not EXCLUDE_NICHE.search(s):
                    results.append(s)
            if done % 20 == 0:
                print(f'   ...{done}/{len(tasks)} done, {len(results)} raw suggestions', file=sys.stderr)

    # Dedupe
    return list(set(results))


def mine_long_tail_variants():
    """Generate long-tail variants from existing high-volume keywords."""
    keywords = json.load(open(KEYWORDS_JSON))
    # Pick top 30 by volume for variant expansion
    sorted_kws = sorted(keywords, key=lambda k: -(k.get('volume') or 0))
    new_kws = set()
    for kw in sorted_kws[:30]:
        base = kw['keyword']
        for tpl in LONG_TAIL_TEMPLATES:
            variant = tpl.format(kw=base.lower())
            if not EXCLUDE_NICHE.search(variant):
                new_kws.add(variant)
    return list(new_kws)


def normalize_keyword(s):
    return re.sub(r'\s+', ' ', s.strip().lower())


def mine_all(args):
    """Run all mining strategies and return combined list."""
    print(f'\n=== Mining strategies ===', file=sys.stderr)

    # Strategy 1: Excel
    print(f'1. Reading Master_Keyword_Plan.xlsx...', file=sys.stderr)
    excel_kws = mine_from_excel()
    print(f'   Got {len(excel_kws)} keywords from Excel', file=sys.stderr)

    # Strategy 2: Google Suggest (only if not dry-run)
    suggest_kws = []
    if not args.dry_run:
        print(f'2. Mining Google Suggest for {len(SERVICES)} services x {len(CITIES)} cities...', file=sys.stderr)
        if args.limit:
            print(f'   (limit: {args.limit} keywords, will stop early)', file=sys.stderr)
        suggest_kws = mine_from_suggestions()
        if args.limit and len(suggest_kws) > args.limit:
            suggest_kws = suggest_kws[:args.limit]
        print(f'   Got {len(suggest_kws)} keywords from Suggest', file=sys.stderr)

    # Strategy 3: Long-tail variants
    print(f'3. Generating long-tail variants...', file=sys.stderr)
    variant_kws = mine_long_tail_variants()
    print(f'   Got {len(variant_kws)} variants', file=sys.stderr)

    # Combine
    all_kws = excel_kws + suggest_kws + variant_kws
    print(f'\nTotal raw keywords: {len(all_kws)}', file=sys.stderr)

    # Normalize + dedupe
    seen = set()
    unique = []
    for kw in all_kws:
        norm = normalize_keyword(kw)
        if norm and norm not in seen:
            seen.add(norm)
            unique.append(kw.strip())

    print(f'Unique after normalize: {len(unique)}', file=sys.stderr)
    return unique


def identify_gaps(unique_kws, slug_set):
    """Filter to keywords that don't yet have a post."""
    gaps = []
    for kw in unique_kws:
        slug = keyword_to_slug(kw)
        if slug in slug_set:
            continue
        # Even if exact slug exists, the keyword might still differ; mark for review
        gaps.append({
            'keyword': kw,
            'keyword_normalized': normalize_keyword(kw),
            'slug': slug,
            'has_post': slug in slug_set,
            'priority_score': priority_score(kw),
            'source': 'miner',
        })
    return gaps


def priority_score(kw):
    """Heuristic priority score (0-100). Higher = more valuable."""
    s = 50
    s += 10 if 'jasa' in kw.lower() else 0
    s += 8 if any(c in kw.lower() for c in ['jakarta', 'bandung', 'surabaya']) else 0
    s += 5 if 'terbaik' in kw.lower() or 'profesional' in kw.lower() else 0
    s += 5 if 'murah' in kw.lower() else 0
    s += 3 if any(c in kw.lower() for c in ['umkm', 'toko', 'jualan']) else 0
    s -= 10 if len(kw) > 60 else 0
    s -= 5 if len(kw) < 10 else 0
    return min(100, max(0, s))


def main():
    p = argparse.ArgumentParser(description='Beriklan.co.id keyword miner')
    p.add_argument('--limit', type=int, help='Limit keywords from Suggest strategy')
    p.add_argument('--all', action='store_true', help='Mine all (no Suggest limit)')
    p.add_argument('--dry-run', action='store_true', help='Preview only, do not write')
    p.add_argument('--no-suggest', action='store_true', help='Skip Google Suggest (use Excel + variants only)')
    args = p.parse_args()

    print(f'\nLoading existing data...', file=sys.stderr)
    existing_kws, slug_set = load_existing_keywords()
    existing_normalized = {normalize_keyword(k['keyword']) for k in existing_kws}

    unique = mine_all(args)

    # Filter out keywords we already have
    new_only = [kw for kw in unique if normalize_keyword(kw) not in existing_normalized]
    print(f'\nNew (not in keywords.json): {len(new_only)}', file=sys.stderr)

    # Build queue entries
    queue = identify_gaps(new_only, slug_set)
    queue.sort(key=lambda x: -x['priority_score'])

    # Attach status
    for i, item in enumerate(queue):
        item['status'] = 'pending'
        item['rank'] = i
        item['created_at'] = datetime.now(timezone.utc).isoformat()

    # Stats
    stats = {
        'last_run': datetime.now(timezone.utc).isoformat(),
        'total_unique_keywords': len(unique),
        'existing_keywords': len(existing_kws),
        'new_keywords': len(new_only),
        'queue_size': len(queue),
        'high_priority': sum(1 for q in queue if q['priority_score'] >= 60),
        'with_existing_post': sum(1 for q in queue if q['has_post']),
        'no_post_yet': sum(1 for q in queue if not q['has_post']),
        'sample_top10': [{'keyword': q['keyword'], 'score': q['priority_score']} for q in queue[:10]],
    }

    print(f'\n=== Queue summary ===', file=sys.stderr)
    print(f'Queue size: {stats["queue_size"]}', file=sys.stderr)
    print(f'High priority (>=60): {stats["high_priority"]}', file=sys.stderr)
    print(f'Without existing post: {stats["no_post_yet"]}', file=sys.stderr)
    print(f'\nTop 10 keywords:', file=sys.stderr)
    for s in stats['sample_top10']:
        print(f'  [{s["score"]:3d}] {s["keyword"]}', file=sys.stderr)

    if args.dry_run:
        print('\n[DRY RUN] Not writing files.', file=sys.stderr)
        return

    os.makedirs(os.path.dirname(QUEUE_OUT), exist_ok=True)
    with open(QUEUE_OUT, 'w') as f:
        json.dump(queue, f, ensure_ascii=False, indent=2)
    with open(STATS_OUT, 'w') as f:
        json.dump(stats, f, ensure_ascii=False, indent=2)
    print(f'\nWrote {QUEUE_OUT}', file=sys.stderr)
    print(f'Wrote {STATS_OUT}', file=sys.stderr)


if __name__ == '__main__':
    main()
